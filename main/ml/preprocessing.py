import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from fuzzywuzzy import fuzz



import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from fuzzywuzzy import fuzz

class ExcelProcessor:
    def __init__(self, file_list):
        self.file_list = file_list
        self.combined_df = pd.DataFrame()
    
    def limiti(self, file_path):
        wb = load_workbook(file_path, data_only=False)
        sheet = wb.active
        limit = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value and str(cell.value).startswith('=SUM'):
                    limit = cell.row
            if limit > 0:
                break
        return limit

    def get_cell_color(self, cell):
        fill = cell.fill
        if fill.start_color.index == '00000000':
            return None
        color = fill.start_color.rgb if fill.start_color.index == 'FFFFFF00' else fill.start_color.index
        return None if color is None or color == '00000000' else str(color)

    def process_file(self, file_path):
        wb = load_workbook(file_path, data_only=False)
        sheet = wb.active
        limit = self.limiti(file_path)
        
        lst_customer, lst_name, lst_city, lst_product_type, lst_count = [], [], [], [], []
        colored_cells = {}
        
        for col_index, col in enumerate(sheet.iter_cols(min_row=3, max_row=limit-1, min_col=1, max_col=sheet.max_column), start=0):
            if col_index % 3 == 0:
                for cell in col:
                    if cell.value and 'נגלה' not in cell.value:
                        header_value = sheet.cell(row=1, column=cell.column).value
                        lst_name.append(header_value)
                        lst_customer.append(cell.value)
        
        for col_index, col in enumerate(sheet.iter_cols(min_row=3, max_row=limit-1, min_col=2, max_col=sheet.max_column), start=0):
            if col_index % 3 == 0:
                for cell in col:
                    if cell.value and 'נגלה' not in cell.value:
                        lst_city.append(cell.value)
        
        for row in sheet.iter_rows(min_row=limit+1, max_row=sheet.max_row, min_col=1, max_col=14):
            for cell in row:
                color = self.get_cell_color(cell)
                if color:
                    left_cell_value = sheet.cell(row=cell.row, column=cell.column-1).value
                    colored_cells[color] = left_cell_value

        for col_index, col in enumerate(sheet.iter_cols(min_row=3, max_row=limit-1, min_col=3, max_col=sheet.max_column), start=0):
            if col_index % 3 == 0:
                for cell in col:
                    if (sheet.cell(row=cell.row, column=cell.column - 1).value is not None) and \
                       (sheet.cell(row=cell.row, column=cell.column - 2).value is not None) and ('נגלה' not in str(cell.value)):
                        color = self.get_cell_color(cell)
                        lst_product_type.append(colored_cells.get(color, 'none'))
                        lst_count.append(cell.value)
        
        min_length = min(len(lst_name), len(lst_customer), len(lst_city), len(lst_product_type), len(lst_count))
        data = {
            'driver_name': lst_name[:min_length],
            'customer': lst_customer[:min_length],
            'city': lst_city[:min_length],
            'product_type': lst_product_type[:min_length],
            'count': lst_count[:min_length]
        }
        df = pd.DataFrame(data)
        filename = os.path.splitext(os.path.basename(file_path))[0]
        try:
            file_date = datetime.strptime(filename.split()[2], '%d.%m.%y')
        except ValueError:
            file_date = datetime.now()
        df['file_date'] = file_date
        return df

    def process_all(self):
        dfs = [self.process_file(file) for file in self.file_list]
        self.combined_df = pd.concat(dfs, ignore_index=True)
        self.clean_data()
        return self.combined_df

    def clean_data(self):
        self.combined_df['count'] = pd.to_numeric(self.combined_df['count'], errors='coerce')
        self.combined_df.dropna(inplace=True)
        
        self.combined_df['customer'] = (
            self.combined_df['customer']
            .str.replace('MGM', '', regex=False)
            .str.replace('OR', '', regex=False)
            .str.strip()
        )
        
        self.combined_df['driver_name'] = self.combined_df['driver_name'].apply(lambda x: 'בא לקחת' if 'בא לקחת' in x else x)
        self.combined_df['city'] = (
            self.combined_df['city']
            .str.replace('מזרכת בתיה', 'מזכרת בתיה', regex=False)
            .str.replace('אלני אבא', 'אלוני אבא', regex=False)
            .str.replace('ביריה', 'טבריה', regex=False)
            .str.replace('א.אל-פחם', 'אום אל-פחם', regex=False)
            .str.strip()
        )
        
        weight_map = {
            'עץ': 20, 'SPC': 23, 'פישבון': 4, 'חבילת פנלים': 10, 'פלור פאן': 12, 'דבק': 15, 'LVT': 15
        }
        self.combined_df['kg'] = self.combined_df['product_type'].map(weight_map).fillna(13)
        
        self.combined_df = self.combined_df[~self.combined_df['driver_name'].isin(['סוואנה', 'בא לקחת', 'תוספת 12/11'])]
        self.combined_df['total_kg'] = self.combined_df['count'] * self.combined_df['kg']

